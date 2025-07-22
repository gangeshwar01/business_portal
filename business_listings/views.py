from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.core.paginator import Paginator
from django.db.models import Q, Avg, Count
from django.http import JsonResponse
from django.views.decorators.http import require_POST
from django.contrib.auth.models import User
from .models import Business, Category, Review, Contact, Newsletter
from .forms import BusinessForm, ReviewForm, ContactForm, NewsletterForm
import json
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import user_passes_test
from .forms import AdminLoginForm
import csv
import io
from django.http import HttpResponse
try:
    import openpyxl
except ImportError:
    openpyxl = None

from django.template.loader import render_to_string
try:
    from fpdf import FPDF
except ImportError:
    FPDF = None
from accounts.models import Notification, SubscriptionPlan, UserSubscription
from django.http import HttpResponseNotFound
from .models import BusinessImage
import warnings
warnings.filterwarnings("ignore", message="Pillow could not be imported - fpdf2 will not be able to add any image")

def home(request):
    """Home page with featured businesses and categories"""
    from .forms import BusinessForm
    featured_businesses = Business.objects.filter(
        status='active', 
        is_featured=True
    ).order_by('-rating', '-created_at')[:6]
    
    categories = Category.objects.annotate(
        business_count=Count('businesses', filter=Q(businesses__status='active'))
    ).order_by('-business_count')[:8]
    
    latest_businesses = Business.objects.filter(
        status='active'
    ).order_by('-created_at')[:8]
    
    top_rated_businesses = Business.objects.filter(
        status='active'
    ).order_by('-rating', '-total_reviews')[:6]
    
    plans = SubscriptionPlan.objects.filter(is_active=True).order_by('price')

    admin_users = User.objects.filter(is_staff=True)
    from .models import Review

    if request.user.is_authenticated:
        # Only show this user's approved businesses
        total_businesses = Business.objects.filter(owner=request.user, status='active').count()
        # Only show reviews made by admin users on this user's businesses
        user_businesses = Business.objects.filter(owner=request.user)
        total_reviews = Review.objects.filter(business__in=user_businesses, user__in=admin_users).count()
    else:
        # Site-wide stats (current logic)
        total_businesses = Business.objects.filter(status='active').count()
        total_reviews = Review.objects.filter(user__in=admin_users).count()

    total_categories = len(BusinessForm.CATEGORY_SPECIFIC_CHOICES)

    context = {
        'featured_businesses': featured_businesses,
        'categories': categories,
        'latest_businesses': latest_businesses,
        'top_rated_businesses': top_rated_businesses,
        'plans': plans,
        'total_businesses': total_businesses,
        'total_reviews': total_reviews,
        'total_categories': total_categories,
    }
    return render(request, 'business_listings/home.html', context)

@login_required
def business_list(request):
    """List all businesses with filtering and search, but only for the current user"""
    businesses = Business.objects.filter(owner=request.user)
    
    # Search functionality
    query = request.GET.get('q')
    if query:
        businesses = businesses.filter(
            Q(name__icontains=query) |
            Q(description__icontains=query) |
            Q(city__icontains=query) |
            Q(state__icontains=query) |
            Q(category__name__icontains=query) |
            Q(address__icontains=query) |
            Q(services__icontains=query) |
            Q(owner_name__icontains=query) |
            Q(owner_email__icontains=query) |
            Q(phone__icontains=query) |
            Q(email__icontains=query)
        )
        # Exclude businesses where all searched fields are None or empty
        businesses = businesses.exclude(
            name__isnull=True, description__isnull=True, city__isnull=True, state__isnull=True,
            category__isnull=True, address__isnull=True, services__isnull=True, owner_name__isnull=True,
            owner_email__isnull=True, phone__isnull=True, email__isnull=True
        )
    
    # Category filter
    category_id = request.GET.get('category')
    if category_id:
        businesses = businesses.filter(category_id=category_id)
    
    # Location filter
    city = request.GET.get('city')
    if city:
        businesses = businesses.filter(city__icontains=city)
    
    # Sort options
    sort_by = request.GET.get('sort', 'newest')
    if sort_by == 'rating':
        businesses = businesses.order_by('-rating', '-total_reviews')
    elif sort_by == 'name':
        businesses = businesses.order_by('name')
    elif sort_by == 'oldest':
        businesses = businesses.order_by('created_at')
    else:  # newest
        businesses = businesses.order_by('-created_at')
    
    # Pagination
    paginator = Paginator(businesses, 12)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    categories = Category.objects.all()
    
    context = {
        'page_obj': page_obj,
        'categories': categories,
        'query': query,
        'selected_category': category_id,
        'selected_city': city,
        'sort_by': sort_by,
        'status': request.GET.get('status'), # Pass status to template
        'admin_review': request.GET.get('admin_review'), # Pass admin_review to template
    }
    return render(request, 'business_listings/business_list.html', context)

@login_required
def business_detail(request, business_id):
    """Detailed view of a business. Owners see their own, staff can see any."""
    if request.user.is_staff:
        business = get_object_or_404(Business, id=business_id)
    else:
        business = get_object_or_404(Business, id=business_id, owner=request.user)
    # Increment view count
    business.views += 1
    business.save()
    # Get reviews
    reviews = business.reviews.all()
    # Check if user has already reviewed
    user_review = None
    if request.user.is_authenticated:
        user_review = reviews.filter(user=request.user).first()
    # Related businesses
    related_businesses = Business.objects.filter(
        category=business.category,
        status='active'
    ).exclude(id=business.id).order_by('-rating')[:4]
    context = {
        'business': business,
        'reviews': reviews,
        'user_review': user_review,
        'related_businesses': related_businesses,
        'admin_review': business.admin_review,
    }
    return render(request, 'business_listings/business_detail.html', context)

@login_required
def add_business(request):
    """Add a new business listing"""
    if request.method == 'POST':
        form = BusinessForm(request.POST, request.FILES)
        
        if form.is_valid():
            business = form.save(commit=False)
            business.owner = request.user
            business.save()
            form.save_m2m()
            
            # Handle gallery images separately
            gallery_files = request.FILES.getlist('gallery_image')
            for img in gallery_files:
                BusinessImage.objects.create(business=business, image=img)
            
            # Notify all admins about new business submission
            from django.contrib.auth.models import User
            admin_users = User.objects.filter(is_staff=True)
            for admin in admin_users:
                Notification.objects.create(
                    user=admin, 
                    message=f'New business "{business.name}" submitted by {request.user.username} for review.'
                )
            
            messages.success(request, 'Business listing submitted successfully! It will be reviewed and activated soon.')
            return redirect('business_listings:business_detail', business_id=business.id)
        else:
            messages.error(request, 'Please correct the errors below.')
    else:
        form = BusinessForm()
    context = {
        'form': form,
        'title': 'Add Business Listing'
    }
    return render(request, 'business_listings/business_form.html', context)

@login_required
def edit_business(request, business_id):
    """Edit an existing business listing"""
    business = get_object_or_404(Business, id=business_id, owner=request.user)
    if request.method == 'POST':
        form = BusinessForm(request.POST, request.FILES, instance=business)
        if form.is_valid():
            business = form.save()
            # Save new images to BusinessImage model
            for img in request.FILES.getlist('gallery_image'):
                BusinessImage.objects.create(business=business, image=img)
            messages.success(request, 'Business listing updated successfully!')
            return redirect('business_listings:business_detail', business_id=business.id)
    else:
        form = BusinessForm(instance=business)
    context = {
        'form': form,
        'business': business,
        'title': 'Edit Business Listing'
    }
    return render(request, 'business_listings/business_form.html', context)

@login_required
@user_passes_test(lambda u: u.is_staff)
def add_review(request, business_id):
    """Add a review for a business (Admin only)"""
    business = get_object_or_404(Business, id=business_id)
    
    # Check if user already reviewed
    existing_review = Review.objects.filter(business=business, user=request.user).first()
    if existing_review:
        messages.warning(request, 'You have already reviewed this business.')
        return redirect('business_listings:business_detail', business_id=business.id)
    
    if request.method == 'POST':
        form = ReviewForm(request.POST)
        if form.is_valid():
            review = form.save(commit=False)
            review.business = business
            review.user = request.user
            review.save()
            messages.success(request, 'Review submitted successfully!')
            return redirect('business_listings:business_detail', business_id=business.id)
    else:
        form = ReviewForm()
    
    context = {
        'form': form,
        'business': business,
        'title': 'Add Review'
    }
    return render(request, 'business_listings/review_form.html', context)

@login_required
def contact(request):
    """Contact form"""
    if request.method == 'POST':
        form = ContactForm(request.POST)
        if form.is_valid():
            contact_obj = form.save()
            # Notify all admins
            from django.contrib.auth.models import User
            from accounts.models import Notification
            admin_users = User.objects.filter(is_staff=True)
            for admin in admin_users:
                Notification.objects.create(user=admin, message=f'New contact message from {contact_obj.name}: {contact_obj.subject}')
            messages.success(request, 'Thank you for your message! We will get back to you soon.')
            return redirect('business_listings:contact')
    else:
        form = ContactForm()
    
    context = {
        'form': form,
        'title': 'Contact Us'
    }
    return render(request, 'business_listings/contact.html', context)

def about(request):
    """About page"""
    admin_users = User.objects.filter(is_staff=True)
    from .models import Review, Business
    if request.user.is_authenticated:
        total_businesses = Business.objects.filter(owner=request.user, status='active').count()
        user_businesses = Business.objects.filter(owner=request.user)
        total_reviews = Review.objects.filter(business__in=user_businesses, user__in=admin_users).count()
    else:
        total_businesses = Business.objects.filter(status='active').count()
        total_reviews = Review.objects.filter(user__in=admin_users).count()
    total_categories = Category.objects.count()
    
    context = {
        'total_businesses': total_businesses,
        'total_reviews': total_reviews,
        'total_categories': total_categories,
        'title': 'About Us'
    }
    return render(request, 'business_listings/about.html', context)

@require_POST
def newsletter_signup(request):
    """Newsletter signup via AJAX"""
    try:
        data = json.loads(request.body)
        email = data.get('email')
        
        if email:
            newsletter, created = Newsletter.objects.get_or_create(email=email)
            if created:
                return JsonResponse({'success': True, 'message': 'Successfully subscribed to newsletter!'})
            else:
                return JsonResponse({'success': True, 'message': 'You are already subscribed!'})
        else:
            return JsonResponse({'success': False, 'message': 'Email is required.'})
    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'message': 'Invalid request.'})

def category_detail(request, category_id):
    """Show businesses in a specific category, only for the owner"""
    if not request.user.is_authenticated:
        return redirect('accounts:login')
    category = get_object_or_404(Category, id=category_id)
    businesses = Business.objects.filter(
        category=category,
        status='active',
        owner=request.user
    ).order_by('-rating', '-created_at')
    
    # Pagination
    paginator = Paginator(businesses, 12)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'category': category,
        'page_obj': page_obj,
        'title': f'{category.name} Businesses'
    }
    return render(request, 'business_listings/category_detail.html', context)

@login_required
def my_businesses(request):
    """User's business listings"""
    businesses = Business.objects.filter(owner=request.user).order_by('-created_at')
    
    context = {
        'businesses': businesses,
        'title': 'My Business Listings'
    }
    return render(request, 'business_listings/my_businesses.html', context)

def search_suggestions(request):
    """AJAX endpoint for search suggestions, only for the owner's businesses"""
    if not request.user.is_authenticated:
        return JsonResponse({'suggestions': []})
    query = request.GET.get('q', '')
    if len(query) < 2:
        return JsonResponse({'suggestions': []})
    businesses = Business.objects.filter(
        Q(name__icontains=query) |
        Q(city__icontains=query) |
        Q(state__icontains=query),
        owner=request.user
    ).filter(status='active')[:5]
    suggestions = []
    for business in businesses:
        suggestions.append({
            'name': business.name,
            'city': business.city,
            'state': business.state,
            'url': business.get_absolute_url()
        })
    return JsonResponse({'suggestions': suggestions})

def admin_login(request):
    if request.user.is_authenticated and request.user.is_staff:
        return redirect('business_listings:admin_dashboard')
    message = ''
    if request.method == 'POST':
        form = AdminLoginForm(request, data=request.POST)
        if form.is_valid():
            user = form.get_user()
            if user.is_staff:
                login(request, user)
                return redirect('business_listings:admin_dashboard')
            else:
                message = 'You do not have admin access.'
        else:
            message = 'Invalid credentials.'
    else:
        form = AdminLoginForm()
    return render(request, 'admin/admin_login.html', {'form': form, 'message': message})

@user_passes_test(lambda u: u.is_staff)
def admin_dashboard(request):
    from .models import Business, Review, Contact
    total_businesses = Business.objects.count()
    pending_businesses = Business.objects.filter(status='pending').count()
    approved_businesses = Business.objects.filter(status='active').count()
    rejected_businesses = Business.objects.filter(status='suspended').count()
    total_reviews = Review.objects.count()
    businesses = Business.objects.all().order_by('-created_at')[:20]
    recent_contacts = Contact.objects.all().order_by('-created_at')[:5]
    pending_subscriptions = UserSubscription.objects.filter(status='pending').order_by('-created_at')
    
    # Get notifications for the current admin user
    admin_notifications = Notification.objects.filter(user=request.user).order_by('-created_at')[:10]
    unread_notification_count = Notification.objects.filter(user=request.user, is_read=False).count()
    
    return render(request, 'admin/admin_dashboard.html', {
        'total_businesses': total_businesses,
        'pending_businesses': pending_businesses,
        'approved_businesses': approved_businesses,
        'rejected_businesses': rejected_businesses,
        'total_reviews': total_reviews,
        'businesses': businesses,
        'recent_contacts': recent_contacts,
        'pending_subscriptions': pending_subscriptions,
        'admin_notifications': admin_notifications,
        'unread_notification_count': unread_notification_count,
    })

@user_passes_test(lambda u: u.is_staff)
@require_POST
def approve_business(request, business_id):
    """Approve a business listing"""
    business = get_object_or_404(Business, id=business_id)
    business.status = 'active'
    business.admin_review = ''  # Clear any previous admin review
    business.save()
    Notification.objects.create(user=business.owner, message=f'Your business "{business.name}" has been approved by admin.')
    messages.success(request, f'Business "{business.name}" has been approved successfully!')
    return redirect('business_listings:admin_dashboard')

@user_passes_test(lambda u: u.is_staff)
@require_POST
def reject_business(request, business_id):
    """Reject a business listing"""
    business = get_object_or_404(Business, id=business_id)
    admin_review = request.POST.get('admin_review', '')
    business.status = 'suspended'
    business.admin_review = admin_review
    business.save()
    Notification.objects.create(user=business.owner, message=f'Your business "{business.name}" has been rejected by admin. Reason: {admin_review}')
    messages.success(request, f'Business "{business.name}" has been rejected.')
    return redirect('business_listings:admin_dashboard')

@user_passes_test(lambda u: u.is_staff)
@require_POST
def delete_business(request, business_id):
    business = get_object_or_404(Business, id=business_id)
    business_name = business.name
    owner = business.owner
    business.delete()
    Notification.objects.create(user=owner, message=f'Your business "{business_name}" has been deleted by admin.')
    messages.success(request, f'Business "{business_name}" has been deleted.')
    return redirect('business_listings:admin_dashboard')

@user_passes_test(lambda u: u.is_staff)
@require_POST
def delete_contact(request, contact_id):
    contact = get_object_or_404(Contact, id=contact_id)
    contact.delete()
    messages.success(request, f'Contact message from {contact.name} has been deleted.')
    return redirect('business_listings:admin_dashboard')

def get_media_urls(business, request):
    """Helper function to get media URLs for a business"""
    media_data = {}
    
    # Handle gallery images
    gallery_urls = []
    for img in business.images.all():
        if img.image:
            gallery_urls.append(request.build_absolute_uri(img.image.url))
    media_data['gallery_images'] = ', '.join(gallery_urls)
    
    # Handle file fields
    file_fields = ['logo', 'cover_image', 'safety_certification_file', 'menu_tariff_card', 'license_certification', 'gst_certification', 'brochure_flyer']
    for field in file_fields:
        file_obj = getattr(business, field, None)
        if file_obj and hasattr(file_obj, 'url'):
            media_data[field] = request.build_absolute_uri(file_obj.url)
        else:
            media_data[field] = ''
    
    return media_data

@user_passes_test(lambda u: u.is_staff)
def export_businesses(request, format):
    from .models import Business
    # Only export businesses owned by the current user unless admin
    if request.user.is_staff:
        businesses = Business.objects.all()
    else:
        businesses = Business.objects.filter(owner=request.user)
    all_fields = [
        'id', 'name', 'category', 'business_types', 'other_business_type', 'year_of_establishment', 'google_business_profile',
        'category_specific_additions', 'room_types_offered', 'room_types_offered_other', 'total_number_of_rooms', 'facilities', 'facilities_other',
        'cuisine_type', 'cuisine_type_other', 'seating_capacity', 'veg_nonveg', 'veg_nonveg_other', 'live_music_pet_friendly', 'live_music_pet_friendly_other',
        'available_services', 'available_services_other',
        'owner_name', 'owner_email', 'owner_area_code', 'owner_phone_number',
        'owner_whatsapp_number', 'owner_alternate_phone_number',
        'description', 'address', 'city', 'state', 'zip_code', 'country', 'phone', 'email', 'website', 'logo', 'cover_image', 'hours_of_operation', 'operating_days', 'services',
        'vehicle_types', 'vehicle_types_other', 'service_area_coverage', 'service_area_coverage_other', 'available_drivers', 'available_drivers_other', 'ac_nonac', 'ac_nonac_other', 'pricing',
        'total_rides_activities', 'entry_fee', 'safety_certification', 'safety_certification_file', 'changing_rooms_lockers',
        'package_types_offered', 'package_types_offered_other', 'destinations_covered', 'price_range', 'transport_accommodation_included', 'transport_accommodation_included_other', 'customization_available', 'special_offers_discounts',
        'menu_tariff_card', 'license_certification', 'gst_certification', 'brochure_flyer',
        'payment_modes_accepted', 'payment_modes_accepted_other', 'ongoing_discounts',
        'declaration_agreed', 'status', 'admin_review', 'created_at', 'updated_at',
        'gallery_images',
    ]
    # Determine which fields have at least one non-empty value
    non_empty_fields = []
    for f in all_fields:
        for b in businesses:
            val = getattr(b, f, '')
            if hasattr(val, 'all'):
                val = ', '.join(str(v) for v in val.all())
            elif isinstance(val, (list, tuple)):
                val = ', '.join(str(v) for v in val)
            elif hasattr(val, 'name'):
                val = str(val.name)
            val = str(val) if val is not None else ''
            if val and val.strip() != '' and val.lower() != 'none':
                non_empty_fields.append(f)
                break
    # CSV Export
    if format == 'csv':
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="businesses.csv"'
        writer = csv.writer(response)
        writer.writerow(non_empty_fields)
        for b in businesses:
            row = []
            media_data = get_media_urls(b, request)
            for f in non_empty_fields:
                if f in media_data:
                    val = media_data[f]
                else:
                    val = getattr(b, f, '')
                    if hasattr(val, 'all'):
                        val = ', '.join(str(v) for v in val.all())
                    elif isinstance(val, (list, tuple)):
                        val = ', '.join(str(v) for v in val)
                    elif hasattr(val, 'name'):
                        val = str(val.name)
                    val = str(val) if val is not None else ''
                row.append(val if val and val.strip() != '' and val.lower() != 'none' else '')
            if any(cell.strip() for cell in row):
                writer.writerow(row)
        return response
    # Excel Export
    elif format == 'excel':
        if not openpyxl:
            return HttpResponse('openpyxl is not installed', status=500)
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.append(non_empty_fields)
        for b in businesses:
            row = []
            media_data = get_media_urls(b, request)
            for f in non_empty_fields:
                if f in media_data:
                    val = media_data[f]
                else:
                    val = getattr(b, f, '')
                    if hasattr(val, 'all'):
                        val = ', '.join(str(v) for v in val.all())
                    elif isinstance(val, (list, tuple)):
                        val = ', '.join(str(v) for v in val)
                    elif hasattr(val, 'name'):
                        val = str(val.name)
                    val = str(val) if val is not None else ''
                row.append(val if val and val.strip() != '' and val.lower() != 'none' else '')
            if any(cell.strip() for cell in row):
                ws.append(row)
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        response = HttpResponse(output, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="businesses.xlsx"'
        return response
    # PDF Export
    elif format == 'pdf':
        if not FPDF:
            return HttpResponse('fpdf2 is not installed', status=500)
        pdf = FPDF(orientation='P', unit='mm', format='A4')
        pdf.set_auto_page_break(auto=True, margin=15)
        for b in businesses:
            pdf.add_page()
            pdf.set_font('Arial', 'B', 14)
            pdf.cell(0, 10, f'Business ID: {b.id}', ln=1)
            pdf.set_font('Arial', '', 10)
            media_data = get_media_urls(b, request)
            for f in non_empty_fields:
                if f in media_data:
                    val = media_data[f]
                else:
                    val = getattr(b, f, '')
                    if hasattr(val, 'all'):
                        val = ', '.join(str(v) for v in val.all())
                    elif isinstance(val, (list, tuple)):
                        val = ', '.join(str(v) for v in val)
                    elif hasattr(val, 'name'):
                        val = str(val.name)
                    val = str(val) if val is not None else ''
                
                if val and val.strip() != '' and val.lower() != 'none':
                    field_name = f.replace('_', ' ').title()
                    pdf.set_font('Arial', 'B', 10)
                    pdf.cell(50, 8, field_name, border=1)
                    pdf.set_font('Arial', '', 10)
                    pdf.cell(0, 8, val[:1000], border=1, ln=1)
        pdf_bytes = pdf.output(dest='S').encode('latin1')
        response = HttpResponse(pdf_bytes, content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="businesses.pdf"'
        return response
    else:
        return HttpResponse('Invalid format', status=400)

def custom_404_view(request, exception=None):
    return render(request, '404.html', status=404)

def privacy_policy(request):
    return render(request, 'business_listings/privacy_policy.html')

def terms_and_conditions(request):
    return render(request, 'business_listings/terms_and_conditions.html')
